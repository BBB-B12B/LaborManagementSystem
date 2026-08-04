from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Shared constants ──
C_HEADER_BG  = "1F3864"
C_HEADER_FG  = "FFFFFF"
C_GRP1_BG    = "D6E4F0"
C_GRP2_BG    = "D5E8D4"
C_GRP3_BG    = "FFF2CC"
C_GRP4_BG    = "FCE4D6"
C_SUBHDR_BG  = "2E75B6"
C_SUBHDR_FG  = "FFFFFF"
C_NOTE_BG    = "F2F2F2"
C_BORDER     = "8EAADB"

ROLES = ["FM", "SE", "LD", "OE", "PE", "PM", "PD", "MD", "AM"]

def thin_border(color=C_BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def cell(ws, row, col, value="", bold=False, fg="000000", bg=None,
         align="left", wrap=False, size=10, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, color=fg, size=size, italic=italic)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    ha = "center" if align == "center" else ("right" if align == "right" else "left")
    c.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=wrap)
    c.border = thin_border()
    return c

# ════════════════════════════════════════════════════════════
# SHEET 1 — Role Guide
# ════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Role Guide"

col_w = [5, 22, 8, 28, 42, 28]
for i, w in enumerate(col_w, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.row_dimensions[1].height = 36
ws1.merge_cells("A1:F1")
c = ws1["A1"]
c.value = "คู่มือการกำหนด Role ในระบบ Labor Management"
c.font = Font(name="Arial", bold=True, color=C_HEADER_FG, size=14)
c.fill = PatternFill("solid", start_color=C_HEADER_BG)
c.alignment = Alignment(horizontal="center", vertical="center")
c.border = thin_border()

ws1.row_dimensions[2].height = 22
ws1.merge_cells("A2:F2")
c = ws1["A2"]
c.value = "เมื่อ Admin เพิ่มสมาชิกใหม่เข้าระบบ ให้เลือก Role ตามตำแหน่งงานจริงของพนักงาน"
c.font = Font(name="Arial", italic=True, color="444444", size=9)
c.fill = PatternFill("solid", start_color="D9E1F2")
c.alignment = Alignment(horizontal="center", vertical="center")
c.border = thin_border()

headers = ["#", "Role (ชื่อตำแหน่ง)", "รหัส", "ใครควรได้ Role นี้", "ทำอะไรได้บ้าง", "หมายเหตุ"]
ws1.row_dimensions[3].height = 22
for col, h in enumerate(headers, 1):
    cell(ws1, 3, col, h, bold=True, fg=C_SUBHDR_FG, bg=C_SUBHDR_BG, align="center", size=10)

rows = [
    ("กลุ่มที่ 1 — ลงรายงานประจำวัน (Daily Report)", None, None, None, None, None, C_GRP1_BG),
    (None, "1", "Foreman", "FM",
     "หัวหน้าคนงานหน้างาน",
     "• ลงรายงานประจำวัน\n• จัดการข้อมูลแรงงานรายวัน (DC)",
     "เห็นเมนู: รายงานประจำวัน, DC เท่านั้น", C_GRP1_BG),
    (None, "2", "Site Engineer", "SE",
     "วิศวกรประจำไซต์",
     "• ลงรายงานประจำวัน\n• ดูภาพรวมงานในไซต์",
     "เห็นเมนู: รายงานประจำวัน เท่านั้น", C_GRP1_BG),
    (None, "3", "Leader", "LD",
     "หัวหน้ากลุ่มงาน",
     "• ลงรายงานประจำวัน\n• ใช้งาน Workspace",
     "เห็นเมนู: รายงานประจำวัน, Workspace", C_GRP1_BG),

    ("กลุ่มที่ 2 — บริหารจัดการโครงการ", None, None, None, None, None, C_GRP2_BG),
    (None, "4", "Office Engineer", "OE",
     "วิศวกรสำนักงาน",
     "• ดู Dashboard\n• สร้าง/จัดการโครงการ\n• ดูค่าแรง\n• ใช้งาน Workspace",
     "ไม่สามารถแก้ไขข้อมูลสมาชิกได้", C_GRP2_BG),
    (None, "5", "Project Engineer", "PE",
     "วิศวกรโครงการ",
     "• ดู Dashboard\n• สร้าง/จัดการโครงการ\n• ดูรายงาน, ค่าแรง\n• ใช้งาน Workspace",
     "ไม่สามารถแก้ไขข้อมูลสมาชิกได้", C_GRP2_BG),
    (None, "6", "Project Manager", "PM",
     "ผู้จัดการโครงการ",
     "• บริหารโครงการที่ได้รับมอบหมาย\n• ดู Dashboard\n• ดูค่าแรง, Workspace",
     "เข้าถึงเฉพาะโครงการที่ได้รับมอบหมาย", C_GRP2_BG),
    (None, "7", "Project Director", "PD",
     "ผู้อำนวยการโครงการ",
     "• บริหารโครงการทุกโครงการในฝ่ายตัวเอง\n• ดู Dashboard, ค่าแรง\n• จัดการสมาชิก (ดูอย่างเดียว)",
     "เข้าถึงเฉพาะโครงการในฝ่ายที่รับผิดชอบ", C_GRP2_BG),

    ("กลุ่มที่ 3 — ผู้บริหารระดับสูง", None, None, None, None, None, C_GRP3_BG),
    (None, "8", "Managing Director", "MD",
     "กรรมการผู้จัดการ",
     "• ดูข้อมูลทุกโครงการในระบบ\n• ดูรายงาน, Dashboard\n• กำหนดกฎประกันสังคม",
     "เห็นภาพรวมทุกโครงการ ไม่สามารถตั้งค่าระบบได้", C_GRP3_BG),

    ("กลุ่มที่ 4 — ผู้ดูแลระบบ", None, None, None, None, None, C_GRP4_BG),
    (None, "9", "Admin", "AM",
     "เจ้าหน้าที่ IT / ผู้ดูแลระบบ",
     "• เพิ่ม/ลบ/แก้ไขสมาชิกทั้งหมด\n• ตั้งค่าค่าแรง\n• นำเข้าข้อมูล Scan\n• กำหนดกฎประกันสังคม",
     "⚠ ควรมีแค่ 1–2 คนในบริษัท", C_GRP4_BG),
]

r = 4
for entry in rows:
    if entry[1] is None:
        ws1.row_dimensions[r].height = 20
        ws1.merge_cells(f"A{r}:F{r}")
        c = ws1[f"A{r}"]
        c.value = entry[0]
        c.font = Font(name="Arial", bold=True, color="1F3864", size=10)
        c.fill = PatternFill("solid", start_color=entry[6])
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = thin_border("1F3864")
        r += 1
    else:
        _, no, role, code, who, what, note, bg = entry
        ws1.row_dimensions[r].height = 52
        cell(ws1, r, 1, no,   align="center", bg=bg)
        cell(ws1, r, 2, role, bold=True, bg=bg)
        cell(ws1, r, 3, code, bold=True, align="center", bg=bg, fg="1F3864")
        cell(ws1, r, 4, who,  bg=bg, wrap=True)
        cell(ws1, r, 5, what, bg=bg, wrap=True, size=9)
        cell(ws1, r, 6, note, bg=bg, wrap=True, size=9, italic=True, fg="444444")
        r += 1

r += 1
ws1.row_dimensions[r].height = 18
ws1.merge_cells(f"A{r}:F{r}")
c = ws1[f"A{r}"]
c.value = "สรุปการเลือก Role — Admin ใช้เป็นแนวทางเพิ่มสมาชิก"
c.font = Font(name="Arial", bold=True, color=C_SUBHDR_FG, size=10)
c.fill = PatternFill("solid", start_color=C_SUBHDR_BG)
c.alignment = Alignment(horizontal="center", vertical="center")
c.border = thin_border()

summary = [
    ("พนักงานหน้างาน (ลงรายงาน)",  "FM, SE, LD"),
    ("วิศวกร / ผู้จัดการโครงการ",   "OE, PE, PM, PD"),
    ("ผู้บริหารดูภาพรวมทั้งหมด",   "MD"),
    ("ดูแลและตั้งค่าระบบ",           "Admin (AM)"),
]
for desc, codes in summary:
    r += 1
    ws1.row_dimensions[r].height = 20
    ws1.merge_cells(f"A{r}:D{r}")
    c = ws1[f"A{r}"]
    c.value = desc
    c.font = Font(name="Arial", size=10)
    c.fill = PatternFill("solid", start_color=C_NOTE_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    c.border = thin_border()
    ws1.merge_cells(f"E{r}:F{r}")
    c = ws1[f"E{r}"]
    c.value = codes
    c.font = Font(name="Arial", bold=True, size=10, color="1F3864")
    c.fill = PatternFill("solid", start_color=C_NOTE_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()

ws1.freeze_panes = "A4"


# ════════════════════════════════════════════════════════════
# SHEET 2 — Page Access Matrix
# ════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Page Access")

# Pages: (page_name, thai_name, group_label)
PAGES = [
    # group, page_name, thai_name
    ("กลุ่มงานหน้างาน", "Daily Reports",           "รายงานประจำวัน"),
    ("กลุ่มงานหน้างาน", "DC Management",           "จัดการแรงงานรายวัน (DC)"),
    ("บริหาร & วิเคราะห์", "Dashboard",            "Dashboard"),
    ("บริหาร & วิเคราะห์", "Management Hub",       "ศูนย์บริหาร (Management)"),
    ("บริหาร & วิเคราะห์", "Project Management",   "จัดการโครงการ"),
    ("บริหาร & วิเคราะห์", "Member Management",    "จัดการสมาชิก"),
    ("บริหาร & วิเคราะห์", "Wage Calculation",     "คำนวณค่าแรง"),
    ("บริหาร & วิเคราะห์", "Workspace",            "Workspace"),
    ("เฉพาะ Admin/MD",    "Scan Data Monitoring",  "ตรวจสอบข้อมูล Scan"),
    ("เฉพาะ Admin/MD",    "Social Security Rules", "กฎประกันสังคม"),
]

# Access matrix: page_name → set of roles that can access
ACCESS = {
    "Daily Reports":          {"FM", "SE", "LD"},
    "DC Management":          {"FM", "AM"},
    "Dashboard":              {"OE", "PE", "PM", "PD", "MD", "AM"},
    "Management Hub":         {"OE", "PE", "PM", "PD", "MD", "AM"},
    "Project Management":     {"PM", "AM"},
    "Member Management":      {"OE", "PE", "PM", "PD", "MD", "AM"},
    "Wage Calculation":       {"OE", "PE", "PM", "PD", "MD", "AM"},
    "Workspace":              {"OE", "PE", "PM", "PD", "MD", "AM", "LD"},
    "Scan Data Monitoring":   {"AM"},
    "Social Security Rules":  {"MD", "AM"},
}

# Role header colors (match sheet 1 groups)
ROLE_HDR_BG = {
    "FM": "2E75B6", "SE": "2E75B6", "LD": "2E75B6",   # blue — หน้างาน
    "OE": "375623", "PE": "375623", "PM": "375623", "PD": "375623",  # green — โครงการ
    "MD": "7F6000",  # yellow-dark — ผู้บริหาร
    "AM": "843C0C",  # orange-dark — Admin
}

# Column setup: col1=group, col2=page(EN), col3=page(TH), col4+=roles
ws2.column_dimensions["A"].width = 22   # group
ws2.column_dimensions["B"].width = 22   # page EN
ws2.column_dimensions["C"].width = 24   # page TH
for i, _ in enumerate(ROLES, 4):
    ws2.column_dimensions[get_column_letter(i)].width = 7

# ── Title ──
ws2.row_dimensions[1].height = 36
total_cols = 3 + len(ROLES)
ws2.merge_cells(f"A1:{get_column_letter(total_cols)}1")
c = ws2["A1"]
c.value = "ตารางสิทธิ์การเข้าถึง Page ตาม Role"
c.font = Font(name="Arial", bold=True, color=C_HEADER_FG, size=14)
c.fill = PatternFill("solid", start_color=C_HEADER_BG)
c.alignment = Alignment(horizontal="center", vertical="center")
c.border = thin_border()

# ── Sub-title row ──
ws2.row_dimensions[2].height = 18
ws2.merge_cells(f"A2:{get_column_letter(total_cols)}2")
c = ws2["A2"]
c.value = "✓ = มีสิทธิ์เข้าถึง   (ช่องว่าง = ไม่มีสิทธิ์)"
c.font = Font(name="Arial", italic=True, color="444444", size=9)
c.fill = PatternFill("solid", start_color="D9E1F2")
c.alignment = Alignment(horizontal="center", vertical="center")
c.border = thin_border()

# ── Header row ──
ws2.row_dimensions[3].height = 22
cell(ws2, 3, 1, "กลุ่มเมนู",    bold=True, fg=C_SUBHDR_FG, bg=C_SUBHDR_BG, align="center")
cell(ws2, 3, 2, "Page (EN)",    bold=True, fg=C_SUBHDR_FG, bg=C_SUBHDR_BG, align="center")
cell(ws2, 3, 3, "Page (ภาษาไทย)", bold=True, fg=C_SUBHDR_FG, bg=C_SUBHDR_BG, align="center")
for ci, role in enumerate(ROLES, 4):
    c2 = ws2.cell(row=3, column=ci, value=role)
    c2.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    c2.fill = PatternFill("solid", start_color=ROLE_HDR_BG[role])
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.border = thin_border()

# ── Data rows ──
prev_group = None
group_bg_map = {
    "กลุ่มงานหน้างาน":    C_GRP1_BG,
    "บริหาร & วิเคราะห์": C_GRP2_BG,
    "เฉพาะ Admin/MD":      C_GRP4_BG,
}

r = 4
for grp, page_en, page_th in PAGES:
    ws2.row_dimensions[r].height = 20
    row_bg = group_bg_map.get(grp, C_NOTE_BG)

    # Group label — show only on first row of each group
    grp_label = grp if grp != prev_group else ""
    cell(ws2, r, 1, grp_label, bold=(grp != prev_group), bg=row_bg,
         fg="1F3864", size=9, wrap=True)
    prev_group = grp

    cell(ws2, r, 2, page_en, bg=row_bg, size=9)
    cell(ws2, r, 3, page_th, bg=row_bg, size=9)

    allowed = ACCESS.get(page_en, set())
    for ci, role in enumerate(ROLES, 4):
        if role in allowed:
            c2 = ws2.cell(row=r, column=ci, value="✓")
            c2.font = Font(name="Arial", bold=True, color="375623", size=11)
            c2.fill = PatternFill("solid", start_color="E2EFDA")
        else:
            c2 = ws2.cell(row=r, column=ci, value="")
            c2.fill = PatternFill("solid", start_color="F5F5F5")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border = thin_border()
    r += 1

ws2.freeze_panes = "A4"


# ════════════════════════════════════════════════════════════
# SHEET 3 — WH Department Note
# ════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("สังกัด WH")

ws3.column_dimensions["A"].width = 28
ws3.column_dimensions["B"].width = 50
ws3.column_dimensions["C"].width = 30

# Title
ws3.row_dimensions[1].height = 36
ws3.merge_cells("A1:C1")
c = ws3["A1"]
c.value = "ข้อกำหนด Role สำหรับสังกัด WH (คลังสินค้า / Warehouse)"
c.font = Font(name="Arial", bold=True, color=C_HEADER_FG, size=13)
c.fill = PatternFill("solid", start_color=C_HEADER_BG)
c.alignment = Alignment(horizontal="center", vertical="center")
c.border = thin_border()

ws3.row_dimensions[2].height = 18
ws3.merge_cells("A2:C2")
c = ws3["A2"]
c.value = "สังกัด WH ใช้ Role ที่แตกต่างจากสังกัดทั่วไป — ไม่มี Role OE และ PE"
c.font = Font(name="Arial", italic=True, color="843C0C", size=9)
c.fill = PatternFill("solid", start_color="FCE4D6")
c.alignment = Alignment(horizontal="center", vertical="center")
c.border = thin_border()

# Header
ws3.row_dimensions[3].height = 22
for ci, h in enumerate(["Role", "คำอธิบาย", "ใช้งานใน WH ได้?"], 1):
    cell(ws3, 3, ci, h, bold=True, fg=C_SUBHDR_FG, bg=C_SUBHDR_BG, align="center")

wh_roles = [
    ("FM — Foreman",          "หัวหน้าคนงานหน้างาน ลงรายงานประจำวัน",              "✓ ใช้ได้",  "E2EFDA", "375623"),
    ("SE — Site Engineer",    "วิศวกรประจำไซต์ ลงรายงานประจำวัน",                  "✓ ใช้ได้",  "E2EFDA", "375623"),
    ("LD — Leader",           "หัวหน้ากลุ่มงาน ลงรายงาน + Workspace",             "✓ ใช้ได้",  "E2EFDA", "375623"),
    ("PM — Project Manager",  "ผู้จัดการโครงการ บริหารงานในคลัง",                   "✓ ใช้ได้",  "E2EFDA", "375623"),
    ("PD — Project Director", "ผู้อำนวยการโครงการ ดูภาพรวมฝ่าย WH",               "✓ ใช้ได้",  "E2EFDA", "375623"),
    ("MD — Managing Director","กรรมการผู้จัดการ ดูข้อมูลทุกโครงการ",               "✓ ใช้ได้",  "E2EFDA", "375623"),
    ("AM — Admin",            "ผู้ดูแลระบบ ตั้งค่าและจัดการสมาชิก",                "✓ ใช้ได้",  "E2EFDA", "375623"),
    ("OE — Office Engineer",  "⛔ ไม่ใช้ใน WH — ตำแหน่งนี้ไม่มีในสายงานคลังสินค้า", "✗ ไม่ใช้", "FFE0E0", "C00000"),
    ("PE — Project Engineer", "⛔ ไม่ใช้ใน WH — ตำแหน่งนี้ไม่มีในสายงานคลังสินค้า", "✗ ไม่ใช้", "FFE0E0", "C00000"),
]

r = 4
for role_name, desc, status, row_bg, status_fg in wh_roles:
    ws3.row_dimensions[r].height = 28
    cell(ws3, r, 1, role_name, bold=True, bg=row_bg, size=10)
    cell(ws3, r, 2, desc, bg=row_bg, wrap=True, size=9)
    c2 = ws3.cell(row=r, column=3, value=status)
    c2.font = Font(name="Arial", bold=True, color=status_fg, size=10)
    c2.fill = PatternFill("solid", start_color=row_bg)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.border = thin_border()
    r += 1

# Note box
r += 1
ws3.row_dimensions[r].height = 18
ws3.merge_cells(f"A{r}:C{r}")
c = ws3[f"A{r}"]
c.value = "หมายเหตุ"
c.font = Font(name="Arial", bold=True, color=C_SUBHDR_FG, size=10)
c.fill = PatternFill("solid", start_color=C_SUBHDR_BG)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
c.border = thin_border()

notes = [
    "• Admin ที่เพิ่มสมาชิกในสังกัด WH ต้องไม่เลือก Role OE หรือ PE เด็ดขาด",
    "• หากต้องการคนดู Dashboard และค่าแรงใน WH ให้ใช้ Role PM หรือ PD แทน",
    "• Role FM, SE, LD ใช้ได้ทั้งสังกัดทั่วไปและสังกัด WH",
]
for note in notes:
    r += 1
    ws3.row_dimensions[r].height = 20
    ws3.merge_cells(f"A{r}:C{r}")
    c = ws3[f"A{r}"]
    c.value = note
    c.font = Font(name="Arial", size=9, color="444444")
    c.fill = PatternFill("solid", start_color=C_NOTE_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    c.border = thin_border()

ws3.freeze_panes = "A4"

# Save
out = "D:/Labor Management System/Role_Guide_v2.xlsx"
wb.save(out)
print("Saved:", out)
